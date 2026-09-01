import openpyxl
import json

wb = openpyxl.load_workbook('scratch/escala_v3.xlsx', data_only=True)
print("Sheet names:", wb.sheetnames)

for name in wb.sheetnames:
    sheet = wb[name]
    print(f"\n--- Sheet: {name} (Max row: {sheet.max_row}, Max col: {sheet.max_column}) ---")
    for r in range(1, min(15, sheet.max_row + 1)):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(15, sheet.max_column + 1))]
        if any(row_vals):
            print(f"Row {r}: {row_vals}")
