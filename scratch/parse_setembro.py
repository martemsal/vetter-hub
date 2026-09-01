import openpyxl
import json
import datetime

wb = openpyxl.load_workbook('scratch/escala_v3.xlsx', data_only=True)
sheet = wb['setembro_26']

print(f"Max rows: {sheet.max_row}, Max cols: {sheet.max_column}")

# Row 4: Dias da semana
# Row 5: Datas
header_days = []
for c in range(5, 36):
    dia_semana = sheet.cell(row=4, column=c).value
    data_val = sheet.cell(row=5, column=c).value
    if data_val:
        if isinstance(data_val, datetime.datetime):
            date_str = data_val.strftime('%d/%m/%Y')
            day_num = str(data_val.day)
        else:
            date_str = str(data_val)
            day_num = str(c - 4)
        header_days.append({
            'col': c,
            'day': day_num,
            'dayOfWeek': dia_semana or '',
            'dateStr': date_str
        })

print(f"Total days found: {len(header_days)}")
for d in header_days[:5]:
    print(d)

# Ler todas as linhas de colaboradores
collaborators = []
current_sector = "Geral"

for r in range(6, sheet.max_row + 1):
    trab = sheet.cell(row=r, column=1).value
    folga = sheet.cell(row=r, column=2).value
    setor = sheet.cell(row=r, column=3).value
    colaborador = sheet.cell(row=r, column=4).value
    
    if setor:
        current_sector = str(setor).strip()
    
    if not colaborador:
        continue
    
    colaborador_str = str(colaborador).strip()
    if colaborador_str.lower() in ['colaborador', 'setor', 'total']:
        continue

    # Pegar os status de cada dia
    days_status = {}
    for hd in header_days:
        val = sheet.cell(row=r, column=hd['col']).value
        val_str = str(val).strip() if val is not None else "Livre"
        days_status[hd['day']] = val_str

    collaborators.append({
        'name': colaborador_str,
        'sector': current_sector,
        'totalTrab': trab,
        'totalFolga': folga,
        'days': days_status
    })

print(f"\nTotal collaborators found: {len(collaborators)}")
sectors_found = set(c['sector'] for c in collaborators)
print("Sectors found:", sectors_found)

for c in collaborators:
    print(f"[{c['sector']}] {c['name']} (Trab: {c['totalTrab']}, Folga: {c['totalFolga']}) -> Dia 1: {c['days'].get('1')}, Dia 6: {c['days'].get('6')}")

# Gerar JSON para o React
output_json = {
    'month': 'Setembro',
    'year': '2026',
    'days': header_days,
    'sectors': sorted(list(sectors_found)),
    'collaborators': collaborators
}

with open('scratch/escala_setembro_parsed.json', 'w', encoding='utf-8') as f:
    json.dump(output_json, f, ensure_ascii=False, indent=2)

print("\nSaved to scratch/escala_setembro_parsed.json successfully!")
